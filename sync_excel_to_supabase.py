import os
import sys
import re
from pathlib import Path
# Try to load using python-dotenv, fallback to manual parsing if missing
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    env_path = Path(__file__).resolve().parent / ".env"
    if env_path.exists():
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    k, v = line.split("=", 1)
                    os.environ[k.strip()] = v.strip()


SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY")

if not SUPABASE_URL or not SUPABASE_ANON_KEY:
    print("Error: SUPABASE_URL and SUPABASE_ANON_KEY are required in .env file.")
    sys.exit(1)

try:
    from openpyxl import load_workbook
    from supabase import create_client
except ImportError:
    print("Required packages 'openpyxl' or 'supabase' are missing.")
    print("Please run with:")
    print("uv run --with openpyxl --with supabase --with python-dotenv sync_excel_to_supabase.py")
    sys.exit(1)

supabase = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
EXCEL_PATH = Path("../[Original] Yapsu AI Curriculum.xlsx").resolve()

if not EXCEL_PATH.exists():
    # Try current directory too
    EXCEL_PATH = Path("[Original] Yapsu AI Curriculum.xlsx").resolve()

if not EXCEL_PATH.exists():
    print(f"Error: Excel file not found at {EXCEL_PATH}")
    sys.exit(1)

def parse_and_sync():
    print(f"Opening Excel workbook: {EXCEL_PATH} ...")
    wb = load_workbook(EXCEL_PATH, data_only=True)
    
    # Initialize basic languages and native languages
    supabase.table("languages").upsert({"id": "zh", "name": "Chinese"}).execute()
    supabase.table("languages").upsert({"id": "ja", "name": "Japanese"}).execute()
    supabase.table("native_languages").upsert({"id": "en", "name": "English"}).execute()
    supabase.table("native_languages").upsert({"id": "vi", "name": "Vietnamese"}).execute()
    
    # Levels Table seeding
    supabase.table("levels").upsert({"id": "level-1-zh", "language_id": "zh", "level_number": 1}).execute()
    
    # Let's parse 'New Yap! AI (Level 1)'
    sheet_name = 'New Yap! AI (Level 1)'
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found. Seeding with basic data instead.")
        return
        
    ws = wb[sheet_name]
    print(f"Parsing sheet: {sheet_name} ...")
    
    current_lesson_id = None
    section = None # 'vocab', 'sentence', 'grammar'
    
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        
        if not cell_val:
            continue
            
        cell_str = str(cell_val).strip()
        
        # Detect Lesson code
        if cell_str.lower() == 'lesson code':
            val = ws.cell(row=row, column=2).value
            if val:
                current_lesson_id = str(val).strip()
                # e.g., 'CN_L101' -> convert to CN_L101_MICRO or keep it
                # If the sheet L1-xx_MICRO exists, we can treat it as micro, else dialogue
                # Let's default lesson type
                lesson_type = 'micro' if 'MICRO' in current_lesson_id else 'dialogue'
                if not current_lesson_id.endswith('_MICRO') and not current_lesson_id.endswith('dialogue'):
                    # Match name
                    pass
                
                lesson_payload = {
                    "id": current_lesson_id,
                    "level_id": "level-1-zh",
                    "code": current_lesson_id,
                    "lesson_number": int(re.search(r'\d+', current_lesson_id).group()) if re.search(r'\d+', current_lesson_id) else 1,
                    "title": f"Lesson {current_lesson_id}",
                    "goal": "Curriculum Sync",
                    "type": lesson_type
                }
                
                # Fetch title from name row if nearby
                title_val = ws.cell(row=row+4, column=4).value
                if title_val:
                    lesson_payload["title"] = str(title_val).strip()
                    
                goal_val = ws.cell(row=row+5, column=4).value
                if goal_val:
                    lesson_payload["goal"] = str(goal_val).strip()
                    
                supabase.table("lessons").upsert(lesson_payload).execute()
                print(f"Synced Lesson: {current_lesson_id} - {lesson_payload['title']}")
                
        # Detect sections
        elif cell_str.lower() == 'vocabulary':
            section = 'vocab'
            continue
        elif cell_str.lower() == 'sentences':
            section = 'sentence'
            continue
        elif cell_str.lower() == 'grammar':
            section = 'grammar'
            continue
        elif cell_str.lower() == 'code' or cell_str.lower() == 'lesson' or cell_str.startswith('L1-'):
            continue
            
        # Insert items based on section
        if current_lesson_id:
            if section == 'vocab' and ('_V' in cell_str or '_v' in cell_str):
                # Row data
                character = ws.cell(row=row, column=2).value
                pinyin = ws.cell(row=row, column=3).value
                english = ws.cell(row=row, column=4).value
                
                if character:
                    vocab_payload = {
                        "id": cell_str,
                        "lesson_id": current_lesson_id,
                        "character": str(character).strip(),
                        "reading_or_pronunciation": str(pinyin).strip() if pinyin else "",
                        "english": str(english).strip() if english else "",
                        "audio_file": f"vocab_{cell_str.split('_')[-1].lower()}.m4a"
                    }
                    supabase.table("vocabularies").upsert(vocab_payload).execute()
                    
            elif section == 'sentence' and ('_S' in cell_str or '_s' in cell_str):
                content = ws.cell(row=row, column=2).value
                pinyin = ws.cell(row=row, column=3).value
                english = ws.cell(row=row, column=4).value
                
                if content:
                    sentence_payload = {
                        "id": cell_str,
                        "lesson_id": current_lesson_id,
                        "content": str(content).strip(),
                        "reading_or_pronunciation": str(pinyin).strip() if pinyin else "",
                        "english": str(english).strip() if english else "",
                        "audio_file": f"sentence_{cell_str.split('_')[-1].lower()}.m4a"
                    }
                    supabase.table("sentences").upsert(sentence_payload).execute()
                    
            elif section == 'grammar' and ('_G' in cell_str or '_g' in cell_str):
                # Grammar info card
                english_desc = ws.cell(row=row, column=4).value
                if english_desc:
                    # Sync dialogue script cards for tutor instructions or grammar tips
                    dialogue_payload = {
                        "lesson_id": current_lesson_id,
                        "segment_id": cell_str,
                        "character_or_role": "Tutor",
                        "text": str(english_desc).strip(),
                        "tts_tag": str(english_desc).strip(),
                        "audio_file": ""
                    }
                    supabase.table("dialogues").upsert(dialogue_payload).execute()

    print("\nSUCCESS: Excel curriculum data synced to Supabase!")

if __name__ == "__main__":
    parse_and_sync()
