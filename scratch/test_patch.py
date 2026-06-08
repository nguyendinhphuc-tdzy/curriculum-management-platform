import sys
from pathlib import Path

# Monkeypatch openpyxl colors
try:
    import openpyxl.styles.colors
    original_rgb_set = openpyxl.styles.colors.RGB.__set__
    def patched_rgb_set(self, instance, value):
        try:
            original_rgb_set(self, instance, value)
        except ValueError:
            # Fallback to transparent/default color if regex check fails
            super(openpyxl.styles.colors.RGB, self).__set__(instance, "00000000")
    openpyxl.styles.colors.RGB.__set__ = patched_rgb_set
    print("Monkeypatch successful!")
except Exception as e:
    print(f"Warning: Failed to monkeypatch openpyxl colors: {e}")

from openpyxl import load_workbook
EXCEL_PATH = Path("../[Original] Yapsu AI Curriculum.xlsx").resolve()
if not EXCEL_PATH.exists():
    EXCEL_PATH = Path("[Original] Yapsu AI Curriculum.xlsx").resolve()

print(f"Attempting to load: {EXCEL_PATH}")
try:
    wb = load_workbook(EXCEL_PATH, data_only=True)
    print("Workbook loaded successfully!")
    print("Available sheets:", wb.sheetnames)
except Exception as e:
    print(f"Failed to load workbook: {e}")
    sys.exit(1)
